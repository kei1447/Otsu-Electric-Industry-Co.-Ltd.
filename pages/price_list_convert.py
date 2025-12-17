import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# --- 設定 ---
# 出力するヘッダー項目（VBAの指定通り）
NEW_COLUMNS = [
    "品目コード", "機種", "品名", "背番号", "単位", 
    "生産原価", "販売単価", "処理平米", "単位重量", "科目名"
]
TARGET_FONT = "游ゴシック"
FONT_SIZE = 11
ROW_HEIGHT = 18.75
COL_WIDTH = 8.38 # 標準より少し狭め（VBA指定値）

def process_excel(file_bytes):
    """
    Excelデータを読み込み、VBAと同様の整形処理を行ってバイナリデータを返す
    """
    # 1. データ読み込み
    # VBAでは「1-2行目を削除」し、「3行目(新1行目)をヘッダーで上書き」していました。
    # つまり、実データは「元の4行目以降」にあると推測されます。
    # そのため、先頭3行(0,1,2)をスキップして読み込みます。
    df = pd.read_excel(file_bytes, header=None, skiprows=3)

    # 列数が足りない/多い場合の安全策（A-Jの10列に絞る）
    df = df.iloc[:, :10]

    # ヘッダーを設定
    df.columns = NEW_COLUMNS

    # --- Excelファイルとしての書き出し処理 (OpenPyXL使用) ---
    output = io.BytesIO()
    
    # Pandasで一旦Excel化（フォーマットなし）
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # シート名は「単価マスタ」
        sheet_name = '単価マスタ'
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        # openpyxlのワークブック・ワークシートオブジェクトを取得
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # 2. 書式設定（VBA: ws.Cells.Font...）
        standard_font = Font(name=TARGET_FONT, size=FONT_SIZE)
        
        # 全セルのフォント設定と行の高さ設定
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = standard_font
            
            # 行の高さ (VBA: RowHeight = 18.75)
            # openpyxlでは行ごとに設定が必要
            worksheet.row_dimensions[row[0].row].height = ROW_HEIGHT

        # 3. 列幅の設定 (VBA: ColumnWidth = 8.38)
        # 全列に対して設定
        for i in range(1, len(NEW_COLUMNS) + 1):
            col_letter = get_column_letter(i)
            worksheet.column_dimensions[col_letter].width = COL_WIDTH

        # 4. テーブルへの変換 (VBA: ListObjects.Add)
        # テーブルの範囲を定義 (A1 : J最終行)
        min_col = 1
        max_col = len(NEW_COLUMNS)
        min_row = 1
        max_row = len(df) + 1 # ヘッダー分+1

        ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        
        tab = Table(displayName="Table1", ref=ref)
        
        # テーブルスタイル（Excel標準の青いやつ）
        style = TableStyleInfo(
            name="TableStyleMedium2", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

    return output.getvalue()

# --- メイン画面 ---
def main():
    st.set_page_config(page_title="Data Converter", layout="wide")
    
    st.title("🏭 電脳工場データ整形ツール")
    st.markdown("""
    電脳工場v1.0から出力された製品リストをアップロードしてください。  
    不要な行の削除、ヘッダー修正、テーブル変換を自動で行います。
    """)

    uploaded_file = st.file_uploader("Excelファイルをアップロード", type=["xlsx", "xls", "xlsm"])

    if uploaded_file:
        if st.button("変換実行"):
            try:
                processed_data = process_excel(uploaded_file)
                
                st.success("変換が完了しました！")
                
                # ダウンロードボタン
                st.download_button(
                    label="整形済みExcelをダウンロード",
                    data=processed_data,
                    file_name=f"formatted_{uploaded_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # プレビュー表示
                st.divider()
                st.caption("▼ 変換後のデータプレビュー")
                # プレビュー用に再度読み込み（フォーマット確認用）
                preview_df = pd.read_excel(io.BytesIO(processed_data))
                st.dataframe(preview_df)

            except Exception as e:
                st.error(f"エラーが発生しました: {e}")

if __name__ == "__main__":
    main()